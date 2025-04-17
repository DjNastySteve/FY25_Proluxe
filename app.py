
import streamlit as st
import pandas as pd
import zipfile
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Proluxe Sales Dashboard", layout="wide")
st.title("📈 Proluxe Sales Dashboard")

@st.cache_data
def load_data():
    sales_df = pd.read_excel("FY25.PLX.xlsx", sheet_name="Sales Data YTD")
    mtd_df = pd.read_excel("FY25.PLX.xlsx", sheet_name="Monthly Goal Sales Data")

    # Rep to name mapping
    cole_reps = ['609', '617', '621', '623', '625', '626']
    jake_reps = ['601', '614', '616', '619', '620', '622', '627']
    "REP": cole_reps + jake_reps + ['Home'],
    "Rep Name": ["Cole"] * len(cole_reps) + ["Jake"] * len(jake_reps) + ["Proluxe"]
})

# Apply Sales Manager filter
if territory != "All":
    df_filtered = df[df["Rep Name"] == territory]
else:
    df_filtered = df.copy()

# Apply Agency filter
if selected_agency != "All":
    df_filtered = df_filtered[df_filtered["Agency"] == selected_agency]


    # Clean columns and parse types
    for df in [sales_df, mtd_df]:
        df.columns = df.columns.str.strip()
        df["Sales Rep"] = df["Sales Rep"].astype(str)
        df["Current Sales"] = pd.to_numeric(df["Current Sales"], errors="coerce").fillna(0)

    return sales_df, mtd_df, rep_map

# Load data once
sales_df, mtd_df, rep_map = load_data()

# Choose MTD or YTD view
view_option = st.sidebar.radio("📅 Select View", ["YTD", "MTD"])
territory = st.sidebar.radio("📌 Select Sales Manager", ["All", "Cole", "Jake", "Proluxe"])
selected_agency = st.sidebar.selectbox("🏢 Filter by Agency", ["All"] + agencies)

df = mtd_df.copy() if view_option == "MTD" else sales_df.copy()
df = df.merge(rep_map, left_on="Sales Rep", right_on="REP", how="left")
# Define agency mapping
rep_agency_mapping = {
    "601": "New Era", "627": "Phoenix", "609": "Morris-Tait", "614": "Access", "616": "Synapse",
    "617": "NuTech", "619": "Connected Sales", "620": "Frontline", "621": "ProAct", "622": "PSG",
    "623": "LK", "625": "Sound-Tech", "626": "Audio Americas"
}
df["Agency"] = df["Sales Rep"].map(rep_agency_mapping)
df["Product Category"] = df["Category 1"]

# Sidebar agency filter
agencies = sorted(df["Agency"].dropna().unique())

# Apply filters

# Banner
if view_option == "MTD":
    banner_html = "<div style='background-color:#212221; padding:1em; border-radius:0.5em; color:#d9d8d6; font-size:18px;'>📅 <b>Now Viewing:</b> <span style='color:#d9d8d6;'>Month-To-Date</span> Performance</div>"
else:
    banner_html = "<div style='background-color:#212221; padding:1em; border-radius:0.5em; color:#d9d8d6; font-size:18px;'>📅 <b>Now Viewing:</b> <span style='color:#d9d8d6;'>Year-To-Date</span> Performance</div>"
st.markdown(banner_html, unsafe_allow_html=True)

# KPIs
total_sales = df_filtered["Current Sales"].sum()
budget = 7538702.63
percent_to_goal = total_sales / budget * 100 if budget > 0 else 0
total_customers = df_filtered["Customer Name"].nunique() if "Customer Name" in df_filtered.columns else 0

col1, col2, col3, col4 = st.columns(4)
col1.metric("📦 Customers", f"{total_customers:,}")
col2.metric("💰 FY25 Sales", f"${total_sales:,.2f}")
col3.metric("🎯 FY25 Budget", f"${budget:,.2f}")
col4.metric("📊 % to Goal", f"{percent_to_goal:.1f}%")

# Export Section
st.sidebar.markdown("### 📤 Export Reports")
rep_options = ["All"] + sorted(df["Sales Rep"].unique().tolist())
selected_export_rep = st.sidebar.selectbox("Choose REP to Export", rep_options)

def generate_dashboard_excel(rep_df, rep):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Rep Dashboard"
    ws["A1"] = "Sales Rep Dashboard"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"], ws["A4"], ws["A5"] = "Sales Rep:", "From:", "To:"
    ws["B3"], ws["B4"], ws["B5"] = rep, "2025-01-01", "2025-12-31"

    sales_by_year = {
        "Goal": 890397.95,
        "Current": rep_df["Current Sales"].sum(),
        "1 Year Ago": 251632,
        "2 Years Ago": 225000,
    }

    for idx, (label, value) in enumerate(sales_by_year.items(), start=7):
        ws.cell(row=idx, column=1, value=label)
        ws.cell(row=idx, column=2, value=value)
        ws.cell(row=idx, column=2).number_format = '"$"#,##0'

    bar_chart = BarChart()
    bar_chart.title = "Sales by Year"
    data = Reference(ws, min_col=2, min_row=6, max_row=10)
    cats = Reference(ws, min_col=1, min_row=7, max_row=10)
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(cats)
    ws.add_chart(bar_chart, "D3")

    ws["F7"], ws["G7"], ws["H7"] = "Current", "Prior", "%"
    ws["F8"] = sales_by_year["Current"]
    ws["G8"] = sales_by_year["1 Year Ago"]
    ws["H8"] = sales_by_year["Current"] / sales_by_year["1 Year Ago"] - 1
    ws["H8"].number_format = "0.0%"

    top_customers = rep_df.groupby("Customer Name")["Current Sales"].sum().sort_values(ascending=False).head(10)
    for idx, (cust, val) in enumerate(top_customers.items(), start=13):
        ws.cell(row=idx, column=5, value=cust)
        ws.cell(row=idx, column=6, value=val)
        ws.cell(row=idx, column=6).number_format = '"$"#,##0'
        ws.cell(row=idx, column=7, value=val / sales_by_year["Current"])
        ws.cell(row=idx, column=7).number_format = "0.0%"

    top_categories = rep_df.groupby("Product Category")["Current Sales"].sum().sort_values(ascending=False).head(5)
    for idx, (cat, val) in enumerate(top_categories.items(), start=25):
        ws.cell(row=idx, column=1, value=cat)
        ws.cell(row=idx, column=2, value=val)
        ws.cell(row=idx, column=2).number_format = '"$"#,##0'

    pie_chart = PieChart()
    pie_chart.title = "Top Product Categories"
    labels = Reference(ws, min_col=1, min_row=25, max_row=25 + len(top_categories) - 1)
    data = Reference(ws, min_col=2, min_row=24, max_row=24 + len(top_categories))
    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(labels)
    ws.add_chart(pie_chart, "D25")

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 15)

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer

if st.sidebar.button("📥 Export Excel Dashboard"):
    if selected_export_rep == "All":
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as zipf:
            for rep in df["Sales Rep"].unique():
                rep_data = df[df["Sales Rep"] == rep]
                buf = generate_dashboard_excel(rep_data, rep)
                zipf.writestr(f"SalesRep_Dashboard_REP{rep}.xlsx", buf.read())
        st.download_button("📦 Download All Dashboards (ZIP)", data=zip_buffer.getvalue(), file_name="All_REPs_Dashboards.zip")
    else:
        rep_data = df[df["Sales Rep"] == selected_export_rep]
        excel_buf = generate_dashboard_excel(rep_data, selected_export_rep)
        st.download_button("📥 Download REP Dashboard", data=excel_buf.getvalue(), file_name=f"SalesRep_Dashboard_REP{selected_export_rep}.xlsx")